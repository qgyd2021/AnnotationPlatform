#!/usr/bin/python3
# -*- coding: utf-8 -*-
import logging
from werkzeug.datastructures import FileStorage

from flask import render_template, request, url_for
import jsonschema
import numpy as np
import openpyxl

from server.exception import ExpectedError
from server.flask_server.route_wrap.common_route_wrap import common_route_wrap
from server.annotation_server.misc.constant import Pages
from server.annotation_server.service.test_tts import BarkTTS
from server.annotation_server.schema import test_tts as test_tts_schema
from server.annotation_server import settings

from toolbox.logging.misc import json_2_str

logger = logging.getLogger('server')


def test_tts():
    kwargs = {
        'bark_tts_by_excel_url': url_for(Pages.bark_tts_by_excel),
    }
    return render_template('test_tts.html', **kwargs)


@common_route_wrap
def bark_tts_view_func():
    args = request.form
    logger.info('bark_tts_view_func, args: {}'.format(json_2_str(args)))

    # request body verification
    try:
        jsonschema.validate(args, test_tts_schema.bark_tts_request_schema)
    except (jsonschema.exceptions.ValidationError,
            jsonschema.exceptions.SchemaError, ) as e:
        raise ExpectedError(
            status_code=60401,
            message='request body invalid. ',
            detail=str(e)
        )

    text = args['text']
    speaker = args['speaker']

    fn_part = BarkTTS.bark_tts(text, speaker)
    result = 'http://{}:{}/{}'.format(settings.bark_tts_host, settings.bark_tts_port, fn_part)
    return result


@common_route_wrap
def bark_speakers_view_func():
    return BarkTTS.bark_speakers()


@common_route_wrap
def bark_tts_by_excel_view_func():
    file_storage: FileStorage = request.files.get('file')
    if not file_storage.filename.endswith('.xlsx'):
        raise ExpectedError(
            status_code=60405,
            message='only excel file supported, filename should endswith .xlsx'
        )

    # read table
    data_excel = openpyxl.load_workbook(file_storage)
    data_sheet = data_excel[data_excel.sheetnames[0]]

    n_rows = data_sheet.max_row
    n_cols = data_sheet.max_column

    row_list = []
    for i in range(1, n_rows + 1):
        row_cells = list()
        for j in range(1, n_cols + 1):
            row_cells.append(data_sheet.cell(i, j).value)
        row_list.append(row_cells)

    # headers verification
    headers = row_list[0]
    try:
        id_idx = headers.index('编号')
        title_idx = headers.index('标题')
        with_audio_idx = headers.index('是否有录音')
        text_idx = headers.index('文案')
    except ValueError as e:
        raise ExpectedError(
            status_code=60405,
            message=str(e)
        )

    data = list()
    for row in row_list[1:]:
        data.append([
            row[id_idx],
            row[title_idx],
            row[with_audio_idx],
            row[text_idx],
        ])

    return 'deprecated'


if __name__ == '__main__':
    pass
